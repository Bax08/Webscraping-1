# Import necessary libraries
import csv
import os
import random
import string
import subprocess
import sys
import time

import requests
import tqdm
from rich import print as rprint
from rich.prompt import Prompt
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.common.desired_capabilities import DesiredCapabilities
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.ui import WebDriverWait
from win32com.client import Dispatch
import openpyxl  # Import the openpyxl library for Excel handling

visit_count = 0

# Function to check internet access
def check_internet_access():
    try:
        response = requests.get("http://www.google.com", timeout=5)
        if response.status_code == 200:
            return True
    except requests.ConnectionError:
        pass
    return False

# Function to get the version of Chrome
def get_chrome_version():
    try:
        parser = Dispatch("Scripting.FileSystemObject")
        paths = [
            r"C:\Program Files\Google\Chrome\Application\chrome.exe",
            r"C:\Program Files (x86)\Google\Chrome\Application\chrome.exe",
        ]
        for path in paths:
            try:
                version = parser.GetFileVersion(path)
                return version
            except Exception:
                pass
    except Exception as e:
        rprint(f"[bold red] Error: {e}[/bold red]")
    return None

# Function to get the version of Chromedriver
def get_chromedriver_version(chromedriver_path):
    try:
        output = subprocess.check_output(
            [chromedriver_path, "--version"], stderr=subprocess.STDOUT, text=True
        )
        version_string = output.strip().split()[1]
        return version_string
    except subprocess.CalledProcessError as e:
        rprint(
            f"[bold red] Error while getting Chromedriver version: {e.output.strip()}[/bold red]"
        )
        return None

# Function to check if the platform is Windows
def is_windows():
    return sys.platform.startswith("win")

# Function to check Chrome and Chromedriver installation and compatibility
def check_chrome_and_chromedriver():
    if not is_windows():
        rprint("[bold red]Error: This script is intended for Windows only.[/bold red]")
        time.sleep(5)
        sys.exit(1)
    else:
        rprint("[bold green]Success: Windows OS Detected![/bold green]")

    if not check_internet_access():
        rprint(
            "[bold red]Error: No internet access. Please make sure you are connected to the internet before running this application![/bold red]"
        )
        time.sleep(5)
        sys.exit(1)

    chrome_paths = [
        r"C:\Program Files\Google\Chrome\Application\chrome.exe",
        r"C:\Program Files (x86)\Google\Chrome\Application\chrome.exe",
    ]
    chrome_installed = any(os.path.exists(path) for path in chrome_paths)
    if not chrome_installed:
        rprint(
            "[bold red]Error: Chrome browser is not installed. Please install Chrome from 'https://www.google.com/chrome/' and try again.[/bold red]"
        )
        time.sleep(5)
        sys.exit(1)

    chromedriver_path = r"C:\chromedriver\chromedriver.exe"
    if not os.path.exists(chromedriver_path):
        rprint(f"Chromedriver not found at: {chromedriver_path}")
        rprint(
            "[bold red]Error: Please download Chromedriver from 'https://chromedriver.chromium.org/downloads' and place it in C:\\chromedriver\\chromedriver.exe[/bold red]"
        )
        time.sleep(5)
        sys.exit(1)

    chromedriver_version = get_chromedriver_version(chromedriver_path)

    if chromedriver_version:
        rprint(f"Chromedriver Version: {chromedriver_version}")
        chrome_version = get_chrome_version()
        if chrome_version:
            rprint(f"Chrome Version: {chrome_version}")

            chromedriver_first_3 = chromedriver_version.split(".")[0:3]
            chrome_first_3 = chrome_version.split(".")[0:3]
            if chromedriver_first_3 == chrome_first_3:
                rprint(
                    "[bold green]Success: Chromedriver and Chrome versions are compatible.[/bold green]"
                )
            else:
                rprint(
                    f"[bold red]Error: Chromedriver(Version {chromedriver_first_3}) and Chrome(Version {chrome_first_3}) versions are not compatible.[/bold red]"
                )
                time.sleep(5)
                sys.exit(1)
        else:
            rprint("[bold red]Error: Could not determine Chrome version.[/bold red]")
            time.sleep(5)
            sys.exit(1)
    else:
        rprint("[bold red]Error: Could not determine Chromedriver version.[/bold red]")
        time.sleep(5)
        sys.exit(1)

# Function to read numbers from Excel file
def read_numbers_from_excel(excel_file):
    numbers = []
    try:
        workbook = openpyxl.load_workbook(excel_file)
        sheet = workbook.active
        for row in sheet.iter_rows(min_row=2, values_only=True):
            numbers.append(row[0])
    except Exception as e:
        rprint(f"[bold red]Error while reading Excel file: {e}[/bold red]")
    return numbers

# Function to generate a random password
def generate_random_password(length=8):
    characters = string.ascii_letters + string.digits + string.punctuation
    return "".join(random.choice(characters) for _ in range(length))

# Function to get free proxies
def get_free_proxies():
    options = webdriver.ChromeOptions()
    options.add_argument("--headless")

    driver = webdriver.Chrome(
        chrome_options=options, executable_path=r"C:\chromedriver\chromedriver.exe"
    )
    driver.get("https://sslproxies.org")

    table = driver.find_element(By.TAG_NAME, "table")
    thead = table.find_element(By.TAG_NAME, "thead").find_elements(By.TAG_NAME, "th")
    tbody = table.find_element(By.TAG_NAME, "tbody").find_elements(By.TAG_NAME, "tr")

    headers = []
    for th in thead:
        headers.append(th.text.strip())

    proxies = []
    for tr in tqdm.tqdm(tbody, desc="Scraping Proxies", colour="green"):
        proxy_data = {}
        tds = tr.find_elements(By.TAG_NAME, "td")
        for i in range(len(headers)):
            proxy_data[headers[i]] = tds[i].text.strip()
        proxies.append(proxy_data)

    driver.quit()

    return proxies

# Function to extract proxy data
def extract_proxy_data():
    free_proxies = get_free_proxies()

    proxy_list = [
        {"IP Address": proxy["IP Address"], "Port": proxy["Port"]}
        for proxy in free_proxies
    ]

    return proxy_list

# Function to update the Excel file with the new status
def update_excel_file(excel_file_path, number_to_update, new_status):
    try:
        workbook = openpyxl.load_workbook(excel_file_path)
        sheet = workbook.active
        found = False
        for row in sheet.iter_rows(min_row=2, values_only=True):
            number, status = row
            if number == number_to_update:
                found = True
                sheet.cell(row=sheet.index(row) + 1, column=2, value=new_status)
            elif not found and not status:
                sheet.cell(row=sheet.index(row) + 1, column=2, value="fail")

        workbook.save(excel_file_path)
        print(f"Status updated to '{new_status}' for number '{number_to_update}' in the Excel file.")
    except Exception as e:
        rprint(f"[bold red]Error while updating Excel file: {e}[/bold red]")

# Function to generate a random 5-digit number
def generate_random_digits():
    return str(random.randint(0, 99999)).zfill(5)

# Function to generate numbers and statuses
def generate_numbers_and_statuses(num_count):
    data = []
    for _ in range(num_count):
        random_digits = generate_random_digits()
        result = "76770" + random_digits
        data.append({"Numbers": result, "Status": ""})
    return data

# Function to create a number Excel file
def create_number_excel_file():
    num_records = 10**5
    excel_file_name = "TEST.xlsx"

    if os.path.exists(excel_file_name):
        options = ["Delete", "Rename", "Keep"]
        choice = Prompt.ask(
            f"The Excel file '{excel_file_name}' already exists. What would you like to do?",
            choices=options,
        )

        if choice == "Delete":
            os.remove(excel_file_name)
        elif choice == "Rename":
            suffix = 1
            while True:
                new_excel_file_name = f"numbers_status_{suffix}.xlsx"
                if not os.path.exists(new_excel_file_name):
                    os.rename(excel_file_name, new_excel_file_name)
                    break
                suffix += 1
            print(f"The Excel file has been renamed to '{new_excel_file_name}'.")
        else:
            print("The existing Excel file will be kept.")
    else:
        print(f"The Excel file '{excel_file_name}' does not exist.")

    data_to_write = generate_numbers_and_statuses(num_records)

    try:
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.append(["Numbers", "Status"])
        for row in data_to_write:
            sheet.append([row["Numbers"], row["Status"]])
        workbook.save(excel_file_name)
        rprint(
            f"[bold green]Success: {num_records} numbers with statuses have been generated and saved to {excel_file_name}.[/bold green]"
        )
    except Exception as e:
        rprint(f"[bold red]Error while creating Excel file: {e}[/bold red]")

# Function to automate the process without a proxy
def automate_without_proxy():
    create_number_excel_file()
    excel_file = "TEST.xlsx"
    numbers = read_numbers_from_excel(excel_file)

    headless_input = input(
        "Run in headless mode?\nNote: Headless mode is a way to run a web browser without a graphical user interface (GUI), making it run in the background without displaying a visible browser window.\n(Yes/No): "
    ).lower()

    if headless_input == "yes":
        headless_mode = True
    else:
        headless_mode = False

    options = webdriver.ChromeOptions()

    if headless_mode:
        options.add_argument("--headless")
    rprint("[bold blue]Initializing automation process![/bold blue]")

    options.add_argument("--disable-extensions")
    options.add_argument("--log-level=3")

    desired_capabilities = DesiredCapabilities.CHROME.copy()
    desired_capabilities["pageLoadStrategy"] = "eager"

    driver = webdriver.Chrome(
        options=options,
        executable_path=r"C:\chromedriver\chromedriver.exe",
        desired_capabilities=desired_capabilities,
    )

    website_url = "https://www.oamfuture.com/index/auth/signup.html"
    for number in numbers:
        driver.get(website_url)

        form_field = driver.find_element(
            By.XPATH, '//*[@id="signup-form"]/div[1]/input'
        )
        form_field.clear()
        form_field.send_keys(number)

        password1 = generate_random_password()
        password2 = password1

        password_field1 = driver.find_element(
            By.XPATH, '//*[@id="signup-form"]/div[2]/input'
        )
        password_field2 = driver.find_element(
            By.XPATH, '//*[@id="signup-form"]/div[3]/input'
        )
        password_field1.send_keys(password1)
        password_field2.send_keys(password2)

        button = driver.find_element(
            By.XPATH, '//*[@id="signup-form"]/div[5]/div/input'
        )
        button.click()

        try:
            text = WebDriverWait(driver, 10).until(
                EC.presence_of_element_located(
                    (By.XPATH, '//div[@class="mui-popup-text"]')
                )
            )

            text_value = text.text.strip()
            if text_value == "success":
                rprint(f"[bold green]Success[/bold green] for Number: {number}")
                update_csv_file(
                    csv_file_path="c:\\Users\\user\\Documents\\upwork\\TEST.xlsx",
                    number_to_update=number,
                    new_status="success",
                )
            elif text_value == "fail":
                rprint(f"[bold red]Fail[/bold red] for Number: {number}")
                update_csv_file(
                    csv_file_path="c:\\Users\\user\\Documents\\upwork\\TEST.xlsx",
                    number_to_update=number,
                    new_status="fail",
                )
            else:
                rprint(f"[bold red]Unexpected text: {text_value}[/bold red]")
                update_csv_file(
                    csv_file_path="c:\\Users\\user\\Documents\\upwork\\TEST.xlsx",
                    number_to_update=number,
                    new_status="fail",
                )
        except:
            rprint(f"[bold red]No text found for Number: {number}[/bold red]")
            update_csv_file(
                csv_file_path="c:\\Users\\user\\Documents\\upwork\\TEST.xlsx",
                number_to_update=number,
                new_status="fail",
            )
        time.sleep(5)
        driver.refresh()

    driver.quit()


def automate_with_proxy():
    """
    The automate_with_proxy function automates the process of signing up for an account on a website.
    It uses a CSV file containing phone numbers to sign up with, and proxies from https://free-proxy-list.net/
    to make each request unique.

    :return: A list of dictionaries
    """
    create_number_excel_file()
    number_csv_file = "c:\\Users\\user\\Documents\\upwork\\TEST.xlsx"
    numbers = read_numbers_from_csv(number_csv_file)

    proxies = extract_proxy_data()
    rprint(f"[bold green]Generated {len(proxies)} Proxy Addresses![/bold green]")

    headless_input = input(
        "Run in headless mode?\nNote: Headless mode is a way to run a web browser without a graphical user interface (GUI), making it run in the background without displaying a visible browser window.\n(Yes/No): "
    ).lower()

    if headless_input == "yes":
        headless_mode = True
    else:
        headless_mode = False

    options = webdriver.ChromeOptions()

    if headless_mode:
        options.add_argument("--headless")

    rprint("[bold blue]Initializing automation process![/bold blue]")

    options.add_argument("--disable-extensions")
    options.add_argument("--log-level=3")

    desired_capabilities = DesiredCapabilities.CHROME.copy()
    desired_capabilities["pageLoadStrategy"] = "eager"

    driver = webdriver.Chrome(
        options=options,
        executable_path=r"C:\chromedriver\chromedriver.exe",
        desired_capabilities=desired_capabilities,
    )

    website_url = "https://www.oamfuture.com/index/auth/signup.html"

    global visit_count

    for number in numbers:
        visit_count += 1

        # Check if it's time to change the proxy
        if visit_count % 5 == 0:
            proxy = random.choice(proxies)
            proxy_address = proxy["IP Address"]
            proxy_port = proxy["Port"]

            proxy_str = f"{proxy_address}:{proxy_port}"
            options.add_argument(f"--proxy-server={proxy_str}")

            driver.quit()  # Close the current driver with the old proxy
            driver = webdriver.Chrome(
                chrome_options=options,
                executable_path=r"C:\chromedriver\chromedriver.exe",
                desired_capabilities=desired_capabilities,
            )

            rprint(
                f"Using Proxy with IP Address: {proxy_address} at Port: {proxy_port}"
            )
        try:
            driver.get(website_url)

            form_field = driver.find_element(
                By.XPATH, '//*[@id="signup-form"]/div[1]/input'
            )
            form_field.clear()
            form_field.send_keys(number)

            password1 = generate_random_password()
            password2 = password1

            password_field1 = driver.find_element(
                By.XPATH, '//*[@id="signup-form"]/div[2]/input'
            )
            password_field2 = driver.find_element(
                By.XPATH, '//*[@id="signup-form"]/div[3]/input'
            )
            password_field1.send_keys(password1)
            password_field2.send_keys(password2)

            button = driver.find_element(
                By.XPATH, '//*[@id="signup-form"]/div[5]/div/input'
            )
            button.click()

            try:
                text = WebDriverWait(driver, 10).until(
                    EC.presence_of_element_located(
                        (By.XPATH, '//div[@class="mui-popup-text"]')
                    )
                )
                text_value = text.text.strip()
                if text_value == "success":
                    rprint(f"[bold green]Success[/bold green] for Number: {number}")
                    update_csv_file(
                        csv_file_path="c:\\Users\\user\\Documents\\upwork\\TEST.xlsx",
                        number_to_update=number,
                        new_status="success",
                    )
                elif text_value == "fail":
                    rprint(f"[bold red]Fail[/bold red] for Number: {number}")
                    update_csv_file(
                        csv_file_path="c:\\Users\\user\\Documents\\upwork\\TEST.xlsx",
                        number_to_update=number,
                        new_status="fail",
                    )
                else:
                    rprint(f"[bold red]Unexpected text: {text_value}[/bold red]")
                    update_csv_file(
                        csv_file_path="c:\\Users\\user\\Documents\\upwork\\TEST.xlsx",
                        number_to_update=number,
                        new_status="fail",
                    )
            except:
                rprint(f"[bold red]No text found for Number: {number}[/bold red]")
                update_csv_file(
                    csv_file_path="c:\\Users\\user\\Documents\\upwork\\TEST.xlsx",
                    number_to_update=number,
                    new_status="fail",
                )
            time.sleep(10)
            driver.refresh()
        except Exception as e:
            rprint(f"Error occured {e}")

    driver.quit()